import pandas as pd
import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog, filedialog
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from collections import defaultdict

# =========================
# 配置
# =========================
BASE_OUTPUT_DIR = "output"

PART_COL = "料号/Part Number"
DESC_COL = "名称/Description"
QTY_COL = "数量/Quantity"
BATCH_COL = "批次/Batch"
EDITION_COL = "版本/Edition"
REMARKS_COL = "备注/Remarks"
NO_COL = "NO."

INFO_FILE = "info.xlsx"
TEMPLATE_FILE = "template.xlsx"

os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)


# =========================
# 工具
# =========================
def clean(v):
    if pd.isna(v): return ""
    return str(v).strip()

def to_int(v):
    """尝试将数量转换为整数，失败则返回0"""
    try:
        return int(float(str(v).strip()))
    except:
        return 0

def clear_template():
    if not os.path.exists(TEMPLATE_FILE): return
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active
    for r in ws.iter_rows(min_row=5):
        for c in r: c.value = None
    wb.save(TEMPLATE_FILE)

def safe_cols(ws):
    cols = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=4, column=c).value
        if v: cols[str(v).strip()] = c
    return cols

# =========================
# 数据逻辑
# =========================
def load_info():
    if not os.path.exists(INFO_FILE): return []
    excel = pd.ExcelFile(INFO_FILE)
    dfs = []
    for sheet in excel.sheet_names:
        df = excel.parse(sheet, dtype=str)
        if PART_COL not in df.columns: continue
        for c in df.columns: df[c] = df[c].apply(clean)
        dfs.append(df)
    return dfs

def search_all(part, dfs):
    part = clean(part)
    results = []
    for df in dfs:
        res = df[df[PART_COL] == part]
        for _, r in res.iterrows():
            results.append({
                "part": clean(r.get(PART_COL)),
                "desc": clean(r.get(DESC_COL)),
                "qty": clean(r.get(QTY_COL)),
                "batch": clean(r.get(BATCH_COL)),
                "edition": clean(r.get(EDITION_COL)),
                "remarks": clean(r.get(REMARKS_COL)),
            })
    return results

def write_items_to_excel(items, target_file):
    wb = load_workbook(target_file)
    ws = wb.active
    cols = safe_cols(ws)
    for r in ws.iter_rows(min_row=5):
        for c in r: c.value = None
    align = Alignment(horizontal="center", vertical="center")
    for i, item in enumerate(items):
        row_idx = 5 + i
        mapping = {NO_COL: i + 1, PART_COL: item["part"], DESC_COL: item["desc"], 
                   QTY_COL: item["qty"], BATCH_COL: item["batch"], 
                   EDITION_COL: item["edition"], REMARKS_COL: item["remarks"]}
        for k, v in mapping.items():
            c_idx = cols.get(k)
            if c_idx:
                cell = ws.cell(row=row_idx, column=c_idx, value=v)
                cell.alignment = align
    wb.save(target_file)

# =========================
# 对比窗口
# =========================
class CompareWin:
    def __init__(self, root, missing, extra):
        win = tk.Toplevel(root)
        win.title("🔍 装箱差异对比报告")
        win.geometry("800x600")

        # 漏装表格
        tk.Label(win, text="❌ 漏装/未装足 (需补货)", fg="red", font=("Arial", 12, "bold")).pack(pady=5)
        self.create_tree(win, missing, ["料号", "缺口数量"]).pack(fill=tk.BOTH, expand=True, padx=10)

        # 多装表格
        tk.Label(win, text="➕ 多装/不在清单内 (需核实)", fg="green", font=("Arial", 12, "bold")).pack(pady=5)
        self.create_tree(win, extra, ["料号", "多出数量"]).pack(fill=tk.BOTH, expand=True, padx=10)

    def create_tree(self, parent, data, headers):
        frame = tk.Frame(parent)
        tree = ttk.Treeview(frame, columns=headers, show="headings", height=8)
        for h in headers:
            tree.heading(h, text=h)
            tree.column(h, anchor="center")
        
        scroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scroll.set)
        
        for item in data:
            tree.insert("", "end", values=item)
            
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        return frame

# =========================
# 主程序
# =========================
class App:
    def __init__(self, root):
        self.root = root
        self.root.title("📦 装箱系统 v2.5 (带对比功能)")
        self.root.geometry("950x600")

        clear_template()
        order = simpledialog.askstring("订单号", "请输入订单号（可空）")

        self.root.deiconify()
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after(100, lambda: self.root.attributes('-topmost', False))
        self.root.focus_force()

        self.output_dir = os.path.join(BASE_OUTPUT_DIR, order) if order else BASE_OUTPUT_DIR
        os.makedirs(self.output_dir, exist_ok=True)

        self.data = load_info()
        self.items = []
        self.current_file = None

        self.input = tk.Entry(root, font=("Arial", 16))
        self.input.pack(fill=tk.X, padx=10, pady=10)
        self.input.bind("<Return>", self.scan)
        self.input.focus_set()

        self.status = tk.Label(root, text=f"📁 新建模式 | {self.output_dir}", fg="blue")
        self.status.pack()

        self.tree = ttk.Treeview(root, columns=("no","part","desc","qty","batch","edition","remarks"), show="headings")
        headers = ["NO.","料号","名称","数量","批次","版本","备注"]
        widths = [60,120,200,80,120,100,200]
        for c, h, w in zip(self.tree["columns"], headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="center")
        self.tree.pack(fill=tk.BOTH, expand=True)

        frame = tk.Frame(root)
        frame.pack(pady=10)

        tk.Button(frame, text="📂 打开已有", width=12, command=self.open_existing).pack(side=tk.LEFT, padx=5)
        self.btn_export = tk.Button(frame, text="📦 生成", width=12, command=self.export)
        self.btn_export.pack(side=tk.LEFT, padx=5)
        self.btn_print = tk.Button(frame, text="🖨 生成并打印", width=14, command=self.export_print)
        self.btn_print.pack(side=tk.LEFT, padx=5)

        self.btn_save_close = tk.Button(frame, text="💾 保存并关闭", width=14, command=self.save_and_close, bg="#e3f2fd")
        
        # 删除按钮
        self.btn_delete = tk.Button(frame, text="❌ 删除选中", width=12, command=self.delete_selected, fg="red")
        self.btn_delete.pack(side=tk.LEFT, padx=5)

        # 对比按钮
        tk.Button(frame, text="📊 差异对比", width=12, command=self.run_comparison, bg="#fff9c4").pack(side=tk.LEFT, padx=5)

    def scan(self, event=None):
        code = self.input.get().strip()
        self.input.delete(0, tk.END)
        res = search_all(code, self.data)
        if not res:
            self.status.config(text=f"❌ 未找到 {code}", fg="red")
            return
        if len(res) == 1: item = res[0]
        else:
            from tkinter import Toplevel
            win = Toplevel(self.root)
            win.title("选择物料")
            tree = ttk.Treeview(win, columns=("part","desc","qty","batch","edition","remarks"), show="headings")
            for h in ["料号","名称","数量","批次","版本","备注"]:
                tree.heading(h, text=h); tree.column(h, width=120, anchor="center")
            tree.pack(fill=tk.BOTH, expand=True)
            for i in res: tree.insert("", "end", values=(i["part"], i["desc"], i["qty"], i["batch"], i["edition"], i["remarks"]))
            res_val = {"v": None}
            def ok():
                sel = tree.selection()
                if sel: res_val["v"] = tree.item(sel[0])["values"]
                win.destroy()
            tk.Button(win, text="确认", command=ok).pack()
            win.grab_set(); self.root.wait_window(win)
            if not res_val["v"]: return
            item = dict(zip(["part","desc","qty","batch","edition","remarks"], res_val["v"]))

        self.items.append(item)
        target = self.current_file if self.current_file else TEMPLATE_FILE
        write_items_to_excel(self.items, target)
        self.tree.insert("", "end", values=(len(self.items), item["part"], item["desc"], item["qty"], item["batch"], item["edition"], item["remarks"]))
        self.status.config(text=f"✅ 已录入 {item['part']}", fg="green")

    def delete_selected(self):
        selected = self.tree.selection()
        if not selected: return
        if not messagebox.askyesno("确认", "确定要删除选中行吗？"): return
        indices = sorted([self.tree.index(item) for item in selected], reverse=True)
        for idx in indices: del self.items[idx]
        for item in selected: self.tree.delete(item)
        for i, child in enumerate(self.tree.get_children()):
            vals = list(self.tree.item(child, 'values'))
            vals[0] = i + 1
            self.tree.item(child, values=vals)
        target = self.current_file if self.current_file else TEMPLATE_FILE
        write_items_to_excel(self.items, target)
        self.status.config(text="⚠️ 已删除并同步文件", fg="orange")

    def run_comparison(self):
        """核心对比功能"""
        # 1. 汇总 Info 需求
        info_total = defaultdict(int)
        for df in self.data:
            for _, row in df.iterrows():
                pn = clean(row.get(PART_COL))
                if pn: info_total[pn] += to_int(row.get(QTY_COL))

        if not info_total:
            messagebox.showerror("错误", "info.xlsx 没有任何有效数据")
            return

        # 2. 选择目录汇总已装箱
        dir_path = filedialog.askdirectory(title="选择要对比的装箱单目录", initialdir=BASE_OUTPUT_DIR)
        if not dir_path: return

        packed_total = defaultdict(int)
        files = [f for f in os.listdir(dir_path) if f.endswith(".xlsx") and f != "info.xlsx" and f != "template.xlsx"]
        
        for f in files:
            try:
                wb = load_workbook(os.path.join(dir_path, f), data_only=True)
                ws = wb.active
                cols = safe_cols(ws)
                pn_col = cols.get(PART_COL)
                qty_col = cols.get(QTY_COL)
                if not pn_col or not qty_col: continue

                r = 5
                while ws.cell(row=r, column=pn_col).value:
                    pn = clean(ws.cell(row=r, column=pn_col).value)
                    qty = to_int(ws.cell(row=r, column=qty_col).value)
                    packed_total[pn] += qty
                    r += 1
            except: continue

        # 3. 比较
        missing = []
        extra = []
        all_parts = set(info_total.keys()) | set(packed_total.keys())

        for p in sorted(all_parts):
            ideal = info_total.get(p, 0)
            actual = packed_total.get(p, 0)
            diff = actual - ideal
            if diff < 0: missing.append((p, abs(diff)))
            elif diff > 0: extra.append((p, diff))

        if not missing and not extra:
            messagebox.showinfo("对比完成", "✨ 完美！所有装箱单与清单完全一致。")
        else:
            CompareWin(self.root, missing, extra)

    def open_existing(self):
        path = filedialog.askopenfilename(filetypes=[("Excel","*.xlsx")])
        if not path: return
        self.current_file = path
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        self.tree.delete(*self.tree.get_children())
        self.items.clear()
        cols = safe_cols(ws)
        pn = cols.get(PART_COL)
        if not pn: return
        r = 5
        while ws.cell(row=r, column=pn).value:
            item = {"part": clean(ws.cell(row=r, column=cols.get(PART_COL)).value),
                    "desc": clean(ws.cell(row=r, column=cols.get(DESC_COL)).value),
                    "qty": clean(ws.cell(row=r, column=cols.get(QTY_COL)).value),
                    "batch": clean(ws.cell(row=r, column=cols.get(BATCH_COL)).value) if cols.get(BATCH_COL) else "",
                    "edition": clean(ws.cell(row=r, column=cols.get(EDITION_COL)).value) if cols.get(EDITION_COL) else "",
                    "remarks": clean(ws.cell(row=r, column=cols.get(REMARKS_COL)).value) if cols.get(REMARKS_COL) else ""}
            self.items.append(item)
            self.tree.insert("", "end", values=(len(self.items), item["part"], item["desc"], item["qty"], item["batch"], item["edition"], item["remarks"]))
            r += 1
        self.status.config(text=f"📄 编辑模式：{os.path.basename(path)}", fg="green")
        self.btn_export.pack_forget(); self.btn_print.pack_forget()
        self.btn_save_close.pack(side=tk.LEFT, padx=5, before=self.btn_delete)

    def save_and_close(self):
        messagebox.showinfo("完成", "更改已保存。")
        self.reset_ui()

    def export(self):
        if not self.items: return
        name = simpledialog.askstring("文件名", "输入名称")
        path = os.path.join(self.output_dir, f"{name if name else 'new'}.xlsx")
        wb = load_workbook(TEMPLATE_FILE); wb.save(path)
        messagebox.showinfo("完成", f"已生成：{path}")
        self.reset_ui()

    def export_print(self):
        if not self.items: return
        name = simpledialog.askstring("文件名", "输入名称")
        path = os.path.join(self.output_dir, f"{name if name else 'new'}.xlsx")
        wb = load_workbook(TEMPLATE_FILE); wb.save(path)
        os.startfile(path, "print")
        messagebox.showinfo("完成", "已打印。")
        self.reset_ui()

    def reset_ui(self):
        self.tree.delete(*self.tree.get_children()); self.items.clear()
        clear_template(); self.current_file = None
        self.status.config(text=f"📁 新建模式 | {self.output_dir}", fg="blue")
        self.btn_save_close.pack_forget()
        self.btn_export.pack(side=tk.LEFT, padx=5, before=self.btn_delete)
        self.btn_print.pack(side=tk.LEFT, padx=5, before=self.btn_delete)

if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()